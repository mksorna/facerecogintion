
from datetime import date     #IMPORTS
from openpyxl.styles import numbers
from openpyxl import load_workbook
import tkinter as tk
import os
from tkinter import messagebox
from cv2 import COLOR_BAYER_BG2RGB, COLOR_BGR2RGB
import mysql.connector 
import cv2,os
import face_recognition
import numpy as np
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidget,QTableWidgetItem
db=mysql.connector.connect(host="localhost",
                        user="root",
                        passwd="root",
                        database="test",
                        )
cursorr=db.cursor(buffered=True)

def formatstrtodate(x):   #STRING DATE TO DATE DATA LIST  yy mm dd
    return [int(i) for i in str(x).split('-')]

def numOfDays(date1, date2):  #FINDS DIFFERENCE BETWEEN TWO DATES
    return (date2-date1).days

class Ui_Dialog(object):#GUI ELEMENTS START HERE
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1368, 700)
        self.label = QtWidgets.QLabel(Dialog)
        self.dep=""     #CONSTRUCTING STUDENT DATA FOR CLASS METHODS
        self.sem=""
        self.idd=""
        self.nameee=""
        self.label.setGeometry(QtCore.QRect(0, 80, 1381, 621))
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap(f"{os.getcwd()}//image//white_back.jpg"))
        self.label.setScaledContents(True)
        self.label.setObjectName("label")
        self.banner = QtWidgets.QLabel(Dialog)
        self.banner.setGeometry(QtCore.QRect(550, 80, 301, 61))
        self.banner.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.banner.setText("")
        self.banner.setPixmap(QtGui.QPixmap(f"{os.getcwd()}//image//boubanner1.png"))
        self.banner.setScaledContents(True)
        self.banner.setObjectName("banner")
        self.title = QtWidgets.QLabel(Dialog)
        self.title.setGeometry(QtCore.QRect(0, 0, 1371, 81))
        self.title.setStyleSheet("font: 75 30pt \"MS Shell Dlg 2\";\n"
"color: rgb(255, 0, 0);\n"
"background-color: rgb(255, 255, 255);")
        self.title.setAlignment(QtCore.Qt.AlignCenter)
        self.title.setObjectName("title")
        self.label_2 = QtWidgets.QLabel(Dialog)
        self.label_2.setGeometry(QtCore.QRect(260, 150, 811, 91))
        self.label_2.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 75 24pt \"MS Shell Dlg 2\";")
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.scan = QtWidgets.QPushButton(Dialog)
        self.scan.setGeometry(QtCore.QRect(570, 490, 211, 41))
        self.scan.setStyleSheet("background-color: rgb(0, 170, 0);\n"
"color: rgb(255, 255, 255);")
        self.scan.setObjectName("scan")
        self.register_2 = QtWidgets.QPushButton(Dialog)
        self.register_2.setGeometry(QtCore.QRect(570, 560, 211, 41))
        self.register_2.setStyleSheet("background-color: rgb(0, 170, 0);\n"
"color: rgb(255, 255, 255);")
        self.register_2.setObjectName("register_2")
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setGeometry(QtCore.QRect(510, 270, 141, 31))
        self.label_3.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(Dialog)
        self.label_4.setGeometry(QtCore.QRect(510, 310, 171, 31))
        self.label_4.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(Dialog)
        self.label_5.setGeometry(QtCore.QRect(510, 350, 141, 31))
        self.label_5.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(Dialog)
        self.label_6.setGeometry(QtCore.QRect(510, 390, 141, 31))
        self.label_6.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(Dialog)
        self.label_7.setGeometry(QtCore.QRect(750, 390, 231, 31))
        self.label_7.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(Dialog)
        self.label_8.setGeometry(QtCore.QRect(750, 310, 231, 31))
        self.label_8.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(Dialog)
        self.label_9.setGeometry(QtCore.QRect(750, 350, 231, 31))
        self.label_9.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(Dialog)
        self.label_10.setGeometry(QtCore.QRect(750, 270, 231, 31))
        self.label_10.setStyleSheet("font: 80 12pt \"MS Shell Dlg 2\";")
        self.label_10.setObjectName("label_10")
        #SIGNALSSIGNALS
        self.scan.clicked.connect(self.scann)
        self.register_2.clicked.connect(self.registerr)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):  #STRINGS AND HTML PARSING BELOW
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.title.setText(_translate("Dialog", "ATTENDANCE SYSTEM USING FACE RECOGNITION"))
        self.label_2.setText(_translate("Dialog", "Focus right on the camera\n"
"You are requested to keep your head steady"))
        self.scan.setText(_translate("Dialog", "Scan Face"))
        self.register_2.setText(_translate("Dialog", "Register attendance"))
        self.label_3.setText(_translate("Dialog", "Department"))
        self.label_4.setText(_translate("Dialog", "Year & Semester"))
        self.label_5.setText(_translate("Dialog", "Student ID"))
        self.label_6.setText(_translate("Dialog", "Student Name"))
        self.label_7.setText(_translate("Dialog", "N/A"))
        self.label_8.setText(_translate("Dialog", "N/A"))
        self.label_9.setText(_translate("Dialog", "N/A"))
        self.label_10.setText(_translate("Dialog", "N/A"))

    def scann(self):   #FACE SCAN
        x=""
        names,ids,encodings_known=[],[],[] #GATHERING NAME DATA,ID DATA,AND ENCODINGS OF ALL IMAGES INSIDE data FOLDER
        for i in os.listdir(f'{os.getcwd()}\\data'):
            sliced=i.split('.')
            names.append(sliced[0].split('_')[1])
            ids.append(int(sliced[0].split('_')[0]))
        for i in os.listdir(f'{os.getcwd()}\\data'):
            data_image = face_recognition.load_image_file(f"{os.getcwd()}\\data\\{i}")
            data_encoding = face_recognition.face_encodings(data_image)[0]
            encodings_known.append(data_encoding)
        video_capture = cv2.VideoCapture(0) #FRAME MONITORING STARTS
        face_locations = []
        face_encodings = []
        face_names = []
        process_this_frame = True
        found=False
        while True:
            # Grab a single frame of video
            ret, frame = video_capture.read() #CAPTURING FRAME

            # Resize frame of video to 1/4 size for faster face recognition processing
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25) 

            # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
            rgb_small_frame = small_frame[:, :, ::-1]

            # Only process every other frame of video to save time
            if process_this_frame:
                # Find all the faces and face encodings in the current frame of video
                face_locations = face_recognition.face_locations(rgb_small_frame)
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

                face_names = []
                for face_encoding in face_encodings:
                    # See if the face is a match for the known face(s)
                    matches = face_recognition.compare_faces(encodings_known, face_encoding,tolerance=0.6)
                    name = "Unknown"
                    if name=="Unknown":
                        found=False
                        x=[("Not found","Not found","Not found","Not found")]
                    

                    # Or instead, use the known face with the smallest distance to the new face
                    face_distances = face_recognition.face_distance(encodings_known, face_encoding)
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        found=True
                        name = names[best_match_index]
                        idd=ids[best_match_index]
                        cursorr.execute("SELECT * FROM systemm WHERE Student_ID=%s",(idd,))
                        x=cursorr.fetchall()
                    face_names.append(name)

            process_this_frame = not process_this_frame


            # Display the results
            for (top, right, bottom, left), name in zip(face_locations, face_names):
                # Scale back up face locations since the frame we detected in was scaled to 1/4 size
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4
                colorrr=(0,255,0) if found else (0, 0, 255)
                # Draw a box around the face
                cv2.rectangle(frame, (left, top), (right, bottom),colorrr , 2)

                # Draw a label with a name below the face
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), colorrr, cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
                
            # Display the resulting image
            cv2.imshow('Video', frame)

            # Hit 'q' on the keyboard to quit!
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        # Release handle to the webcam
        video_capture.release()
        cv2.destroyAllWindows()
        print(x)
        if name=='Unknown':
            self.register_2.setEnabled(False)
        self.label_10.setText(x[0][0])  #Department
        self.label_8.setText(x[0][1]) #Semester
        self.label_9.setText(str(x[0][2])) #ID
        self.label_7.setText(x[0][3]) #Name
        if name!='Unknown':
            self.dep=x[0][0]
            self.sem=x[0][1]
            self.idd=x[0][2]
            self.nameee=x[0][3]
    def registerr(self):  #ENTERS DATA INSIDE EXCEL SHEET
        found=False
        wb= load_workbook(f'{self.dep}.xlsx')
        ws=wb[self.sem]
        today = date.today()
        for i in ws['C']:
            if str(i.value)==str(self.idd):
                
                id_cell=f'C{i.row}'
                found=True
                date_cell=f'A{i.row}'
                status_cell=f'D{i.row}'
                last_update=ws[date_cell].value
                
                formatted_last_update=date(*(formatstrtodate(last_update)))
                difference=numOfDays(formatted_last_update,date(*formatstrtodate(today)))
                if difference==0:
                    break
                elif difference>0:
                   
                    ws[date_cell]=str(today)
                    
                    if ws[status_cell].value==None or ws[status_cell].value=="" or ws[status_cell].value=='Absent':
                        ws[status_cell]='Present'
                        break
        
        if not found:
            if ws[f'A{ws.max_row-1}']==None:
                rc=ws.max_row
            else:
                rc=ws.max_row+1
            print('Suppposed to be saved')
            ws[f"A{rc}"].number_format=numbers.FORMAT_TEXT
            ws[f"A{rc}"],ws[f"B{rc}"],ws[f"C{rc}"],ws[f"D{rc}"]=str(today),self.nameee,str(self.idd),'Present'
            

            #ws.append([str(today),self.nameee,int(self.idd),'Present'])
        wb.save(f'{self.dep}.xlsx')

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())
